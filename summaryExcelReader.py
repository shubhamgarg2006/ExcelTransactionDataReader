import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# --- Automatically use Excel file in the same folder as the script ---
script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, "portfolio.xlsx")

# --- Load Excel ---
start_positions = pd.read_excel(file_path, sheet_name="Positions")
transactions = pd.read_excel(file_path, sheet_name="Transactions")

# --- Initialize portfolio ---
portfolio = {}
for idx, row in start_positions.iterrows():
    ticker = row['Ticker']
    qty = row.get('Quantity', 0) if not pd.isna(row.get('Quantity', 0)) else 0
    cash = row.get('Cash', 0) if not pd.isna(row.get('Cash', 0)) else 0
    if ticker.lower() == 'cash':
        portfolio['Cash'] = cash
    else:
        portfolio[ticker] = qty

if 'Cash' not in portfolio:
    portfolio['Cash'] = 0

# --- Sort and group transactions by date ---
transactions['Date'] = pd.to_datetime(transactions['Date'])
transactions = transactions.sort_values('Date')
grouped = transactions.groupby('Date')

# --- Process each date's transactions cumulatively ---
positions_over_time = []
transaction_summaries = []

for date, day_trades in grouped:
    # Build summary for this date
    summary = {
        'date': date,
        'buys': [],
        'sells': [],
        'dividends': 0,
        'interest': 0,
        'other': 0,
        'total_cash_change': 0
    }
    
    for idx, row in day_trades.iterrows():
        action = row['Action'].lower().strip()
        ticker = row['Ticker']
        qty = row.get('Quantity', 0) if not pd.isna(row.get('Quantity', 0)) else 0
        amount = row.get('Amount', 0) if not pd.isna(row.get('Amount', 0)) else 0
        trans_price = row.get('TransactionPrice', 0) if 'TransactionPrice' in row and not pd.isna(row['TransactionPrice']) else 0

        # Track for summary
        summary['total_cash_change'] += amount
        
        # Ensure ticker exists if stock transaction
        if ticker not in portfolio and action in ['buy', 'sell']:
            portfolio[ticker] = 0

        # --- Apply transaction rules ---
        if action == 'buy':
            portfolio[ticker] = portfolio.get(ticker, 0) + qty
            portfolio['Cash'] += (amount)   # pay cash for stock + transaction fee
            summary['buys'].append(f"{abs(qty)} shares of {ticker} for ${abs(amount):.2f}")
        elif action == 'sell':
            portfolio[ticker] = portfolio.get(ticker, 0) + qty  # qty negative for sells
            portfolio['Cash'] += (amount)  # receive sale proceeds minus transaction fee
            summary['sells'].append(f"{abs(qty)} shares of {ticker} for ${amount:.2f}")
        elif action == 'div':
            portfolio['Cash'] += amount
            summary['dividends'] += amount
        elif action == 'int':
            portfolio['Cash'] += amount
            summary['interest'] += amount
        elif action == 'other':
            portfolio['Cash'] += amount
            summary['other'] += amount

    # --- Remove tickers with zero quantity ---
    tickers_to_remove = [t for t, q in portfolio.items() if t != 'Cash' and q == 0]
    for t in tickers_to_remove:
        del portfolio[t]

    # --- Snapshot after all trades that day ---
    snapshot = {'Date': date}
    snapshot.update(portfolio)
    positions_over_time.append(snapshot)
    transaction_summaries.append(summary)

# --- Write formatted text file with summaries ---
text_output = os.path.join(script_dir, "portfolio_positions.txt")
with open(text_output, "w") as f:
    for i, snapshot in enumerate(positions_over_time):
        summary = transaction_summaries[i]
        
        f.write(f"=" * 60 + "\n")
        f.write(f"Date: {snapshot['Date'].strftime('%Y-%m-%d')}\n")
        f.write(f"-" * 60 + "\n")
        
        # Write transaction summary
        f.write("TRANSACTION SUMMARY:\n")
        if summary['buys']:
            f.write("  Buys:\n")
            for buy in summary['buys']:
                f.write(f"    - {buy}\n")
        if summary['sells']:
            f.write("  Sells:\n")
            for sell in summary['sells']:
                f.write(f"    - {sell}\n")
        if summary['dividends'] != 0:
            f.write(f"  Dividends: ${summary['dividends']:.2f}\n")
        if summary['interest'] != 0:
            f.write(f"  Interest: ${summary['interest']:.2f}\n")
        if summary['other'] != 0:
            f.write(f"  Other: ${summary['other']:.2f}\n")
        f.write(f"  Total Cash Change: ${summary['total_cash_change']:.2f}\n")
        f.write("\n")
        
        # Write positions
        f.write("POSITIONS AFTER TRANSACTIONS:\n")
        f.write(f"  Cash: ${snapshot['Cash']:.2f}\n")
        for ticker, qty in snapshot.items():
            if ticker not in ['Date', 'Cash']:
                f.write(f"  {ticker}: {qty} shares\n")
        f.write("\n")

# --- Write Excel file with summaries ---
excel_rows = []
for i, snapshot in enumerate(positions_over_time):
    summary = transaction_summaries[i]
    date_str = snapshot['Date'].strftime('%Y-%m-%d')
    
    # Add summary section
    excel_rows.append({'Date': date_str, 'Section': 'SUMMARY', 'Field': 'Total Cash Change', 'Value': summary['total_cash_change']})
    
    if summary['buys']:
        for buy in summary['buys']:
            excel_rows.append({'Date': date_str, 'Section': 'SUMMARY', 'Field': 'Buy', 'Value': buy})
    if summary['sells']:
        for sell in summary['sells']:
            excel_rows.append({'Date': date_str, 'Section': 'SUMMARY', 'Field': 'Sell', 'Value': sell})
    if summary['dividends'] != 0:
        excel_rows.append({'Date': date_str, 'Section': 'SUMMARY', 'Field': 'Dividends', 'Value': summary['dividends']})
    if summary['interest'] != 0:
        excel_rows.append({'Date': date_str, 'Section': 'SUMMARY', 'Field': 'Interest', 'Value': summary['interest']})
    if summary['other'] != 0:
        excel_rows.append({'Date': date_str, 'Section': 'SUMMARY', 'Field': 'Other', 'Value': summary['other']})
    
    # Add positions section
    excel_rows.append({'Date': date_str, 'Section': 'POSITIONS', 'Field': 'Cash', 'Value': snapshot['Cash']})
    for ticker, qty in snapshot.items():
        if ticker not in ['Date', 'Cash']:
            excel_rows.append({'Date': date_str, 'Section': 'POSITIONS', 'Field': ticker, 'Value': qty})

excel_output = os.path.join(script_dir, "portfolio_positions_readable.xlsx")
excel_df = pd.DataFrame(excel_rows)
excel_df.to_excel(excel_output, index=False)

# --- Apply conditional formatting ---
wb = load_workbook(excel_output)
ws = wb.active

# Define formatting styles
summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
positions_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue
header_font = Font(bold=True, color="FFFFFF")  # White bold text

# Format header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Format data rows based on Section column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    section_value = row[1].value  # Section is column B (index 1)
    
    if section_value == 'SUMMARY':
        for cell in row:
            cell.fill = summary_fill
    elif section_value == 'POSITIONS':
        for cell in row:
            cell.fill = positions_fill

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(excel_output)

print("✅ Portfolio positions saved to:")
print(f"   - {text_output}")
print(f"   - {excel_output}")
print("\n📊 Excel formatting applied:")
print("   - SUMMARY rows: Light yellow background")
print("   - POSITIONS rows: Light blue background")